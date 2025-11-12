#!/usr/bin/env python3
#
# Copyright 2025 NXP
#
# SPDX-License-Identifier: Apache-2.0

import argparse
import logging
from collections.abc import Iterable
from itertools import chain
from pathlib import Path
from textwrap import dedent

import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from zephyr_repo_tools.bindings import ZephyrDTSBindings
from zephyr_repo_tools.compat import ZephyrDTSCompats
from zephyr_repo_tools.iterfiles import iter_files

green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

center_align = Alignment(horizontal="center", vertical="center")


def configure_logging(verbosity: int) -> None:
    """
    Configure logging level based on verbosity count.
    -v  -> INFO
    -vv -> DEBUG
    """
    if verbosity == 0:
        level = logging.WARNING
    elif verbosity == 1:
        level = logging.INFO
    else:
        level = logging.DEBUG

    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def collect_unique_keys(boards: Iterable["ZephyrDTSCompats"]) -> set[str]:
    """
    Collect the union of all compatible keys across a set of boards.

    Args:
        boards (Iterable[ZephyrDTSCompats]): Iterable of board objects.

    Returns:
        set[str]: A set containing all unique compatible keys.
    """
    return set(chain.from_iterable(b.compatibles.keys() for b in boards))


def to_dataframe(
    boards: Iterable["ZephyrDTSCompats"],
    bindings: "ZephyrDTSBindings",
) -> pd.DataFrame:
    """
    Build a DataFrame summarizing compatible strings across boards.

    Args:
        boards (Iterable[ZephyrDTSCompats]): Iterable of board objects.
        bindings (ZephyrDTSBindings): Bindings object for resolving types.

    Returns:
        pd.DataFrame: A DataFrame with rows for each compatible string and
                      columns for type and board-specific counts.
    """
    boards = list(boards)
    board_names = [Path(getattr(b, "filename", None)).stem or str(b) for b in boards]

    all_compats = collect_unique_keys(boards)

    rows = []
    for compat in all_compats:
        r = {
            "compat": compat,
            "type": bindings.get_type(compat),
        }
        for name, b in zip(board_names, boards, strict=True):
            r[name] = b.compatibles.get(compat, '')
        rows.append(r)

    cols = ["compat", "type"] + board_names
    df = pd.DataFrame(rows, columns=cols)

    # Remove numeric index
    df.reset_index(drop=True, inplace=True)
    return df


def build_board_type_df(
    boards: list["ZephyrDTSCompats"],
    bindings: "ZephyrDTSBindings",
) -> pd.DataFrame:
    """
    Build a DataFrame summarizing board types and their counts.

    Args:
        boards (list[ZephyrDTSCompats]): Sorted list of board objects.
        bindings (ZephyrDTSBindings): Provides `get_type(compat)` for resolving compatibilities.

    Returns:
        pd.DataFrame: A DataFrame with one row per board and columns for each type.
    """
    # Resolve board names
    board_names = [getattr(b, "filename", None) or str(b) for b in boards]

    # Collect all types across all boards
    all_types: set[str] = {
        t
        for b in boards
        for compat in b.compatibles
        if (t := bindings.get_type(compat)) is not None
    }
    type_list = sorted(all_types)

    # Build rows: one per board, summing counts per type
    rows: list[dict[str, int | str]] = []
    for name, b in zip(board_names, boards, strict=True):
        row = {"board": name, **{t: 0 for t in type_list}}
        for compat, cnt in b.compatibles.items():
            if (t := bindings.get_type(compat)) is not None:
                row[t] += int(cnt)
        rows.append(row)

    # Create DataFrame and replace zeros with empty strings
    df = pd.DataFrame(rows, columns=["board"] + type_list)
    for t in type_list:
        df[t] = df[t].replace(0, "")

    return df


# ---------------- openpyxl formatting ----------------
def _format_headers_openpyxl(ws, df: pd.DataFrame, fixed_cols: set[str]) -> None:
    """
    Rotate all headers except those in fixed_cols, autosize header row height,
    and size fixed columns by their content length.
    """
    rotate_cols = [c for c in df.columns if c not in fixed_cols]
    rotated_headers: list[str] = []

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        col_letter = get_column_letter(col_idx)

        if col_name in rotate_cols:
            # Rotate header up; keep column narrow
            cell.alignment = Alignment(
                textRotation=90, horizontal="center", vertical="bottom", wrap_text=True
            )
            ws.column_dimensions[col_letter].width = 4
            rotated_headers.append(str(col_name))
        else:
            # Size based on longest content (header + values)
            series = df[col_name].astype(str)
            max_len = max(series.map(len).max(), len(str(col_name)))
            ws.column_dimensions[col_letter].width = max(12, min(max_len + 2, 80))

    # Auto header row height: scale with font size and longest rotated header
    font_sz = float(ws.cell(row=1, column=1).font.sz or 11)
    longest = max((len(h) for h in rotated_headers), default=0)
    row_height_pts = longest * (0.56 * font_sz) + 0.9 * font_sz  # heuristic; no low clamp
    ws.row_dimensions[1].height = min(row_height_pts, 409.5)  # cap only at Excel max


def export_views_xlsx_openpyxl(
    df_matrix: pd.DataFrame,
    df_board_type: pd.DataFrame,
    path: str,
    sheet1: str = "Matrix",
    sheet2: str = "Board×Type",
) -> None:
    """
    Export two pandas DataFrames to an Excel file using openpyxl with custom formatting.

    Args:
        df_matrix (pd.DataFrame): DataFrame for the first sheet.
        df_board_type (pd.DataFrame): DataFrame for the second sheet.
        path (str): Output Excel file path.
        sheet1 (str, optional): Name of the first sheet. Defaults to "Matrix".
        sheet2 (str, optional): Name of the second sheet. Defaults to "Board×Type".

    Returns:
        None: Writes the Excel file to the specified path.
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Sheet 1: compat × boards
        df_matrix.to_excel(writer, index=False, sheet_name=sheet1)
        ws1 = writer.sheets[sheet1]
        _format_headers_openpyxl(ws1, df_matrix, fixed_cols={"compat", "type"})
        ws1.freeze_panes = "C2"  # freeze top row + first two cols

        for row in range(2, ws1.max_row + 1):  # Start from row 2
            for col in range(3, ws1.max_column + 1):  # Start from column 3
                cell = ws1.cell(row=row, column=col)
                cell.alignment = center_align
                # Apply alternating shading starting from row 2 (since row 1 is header)
                if row % 2 == 0:  # even rows
                    cell.fill = light_gray

        # Sheet 2: board × type
        df_board_type.to_excel(writer, index=False, sheet_name=sheet2)
        ws2 = writer.sheets[sheet2]
        _format_headers_openpyxl(ws2, df_board_type, fixed_cols={"board"})
        ws2.freeze_panes = "B2"  # freeze top row + first column

        for row in range(2, ws2.max_row + 1):
            for col in range(2, ws2.max_column + 1):
                cell = ws2.cell(row=row, column=col)
                cell.alignment = center_align
                # Apply alternating shading starting from row 2 (since row 1 is header)
                if row % 2 == 0:  # even rows
                    cell.fill = light_gray


def main():
    class _Fmt(argparse.ArgumentDefaultsHelpFormatter, argparse.RawTextHelpFormatter):
        pass

    parser = argparse.ArgumentParser(
        prog="zephyr_scrape_boards.py",
        description=(
            "Recursively extract 'compatible' strings from Zephyr DTS files and "
            "optionally export summarized views to Excel."
        ),
        allow_abbrev=False,
        epilog=dedent(
            """\
            Examples:
              # Scan a specific board directory and write an Excel matrix

              zephyr_scrape_boards.py """
            "-I zephyr\\dts\\arm -I zephyr\\dts\\arm64 -I zephyr\\dts "
            "-I zephyr\\dts\\common -I modules\\hal\\nxp\\dts -b zephyr\\dts\\bindings "
            "-x nxp_boards.xlsx zephyr\\boards\\nxp\\"
        ),
        formatter_class=_Fmt,
    )

    parser.add_argument(
        "dts_file",
        type=Path,
        help=(
            "Path to the top-level DTS file or a directory to scan. "
            "If a directory is given, all '*.dts' files are processed recursively."
        ),
    )
    parser.add_argument(
        "-I",
        "--include",
        action="append",
        default=[],
        metavar="PATH",
        help=(
            "Additional include path(s) for resolving #include in DTS. "
            "Can be specified multiple times."
        ),
    )
    parser.add_argument(
        "-b",
        "--bindings",
        type=Path,
        metavar="DIR",
        required=False,
        help="Path to the Zephyr devicetree bindings directory (e.g., zephyr/dts/bindings).",
    )
    parser.add_argument(
        "-x",
        "--xlsx",
        metavar="FILE",
        help="Write results to an Excel workbook (.xlsx). If omitted, results are printed to \
        stdout.",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="count",
        default=0,
        help="Increase log verbosity (-v for INFO, -vv for DEBUG).",
    )
    args = parser.parse_args()

    configure_logging(args.verbose)
    logging.info("Starting DTS processing...")
    logging.debug("Arguments: %s", args)

    dts_list = []
    bindings_collection = ZephyrDTSBindings(args.bindings)

    p = Path(args.dts_file)
    if p.is_file():
        root_dir = p.parent
    else:
        root_dir = p

    # Always add the root_dir as an include path for relative includes to resolve.
    args.include.append(str(root_dir))

    for f in iter_files(args.dts_file, [".dts"]):
        dts_entry = ZephyrDTSCompats(f, args.include, bindings_collection)
        dts_list.append(dts_entry)

    sorted_files = sorted(dts_list)

    if args.xlsx:
        logging.info("Writing XLSX output to %s", args.xlsx)
        df = to_dataframe(sorted_files, bindings_collection)
        df2 = build_board_type_df(sorted_files, bindings_collection)
        export_views_xlsx_openpyxl(df, df2, args.xlsx)
    else:
        logging.info("No output file specified; printing to console")
        for dfile in sorted_files:
            print("\n" + "-" * 78 + "\n")
            print(f"{dfile.filename}:\n")
            for item in sorted(dfile.compatibles):
                print(item)

    logging.info("Processing complete.")


if __name__ == "__main__":
    main()
