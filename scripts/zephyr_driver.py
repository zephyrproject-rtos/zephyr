#!/usr/bin/env python3
#
# Copyright 2025 NXP
#
# SPDX-License-Identifier: Apache-2.0

import argparse
import glob
import logging
import os
from collections import defaultdict
from pathlib import Path

# Third-party
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Local application
from zephyr_repo_tools.api_scan import scan_api_structs
from zephyr_repo_tools.driver_info import ZephyrDriver


def configure_logging(verbosity: int) -> None:
    """
    Configure logging level based on verbosity count.
    -v  -> INFO
    -vv -> DEBUG
    """
    if verbosity == 0:
        level = logging.WARNING
    elif verbosity == 1:
        level = logging.INFO
    else:
        level = logging.DEBUG

    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def find_driver_files(patterns, compat_only=False, recursive=False, zephyr_def=None):
    entries = []

    # If recursive flag is set but pattern doesn't use **, modify patterns
    if recursive:
        patterns = [p.replace('*/', '**/') if '**' not in p else p for p in patterns]

    found_files = set()

    # Process each pattern
    for pattern in patterns:
        # Find matching files
        logging.debug(f"pattern: {pattern}")
        matching_files = glob.glob(pattern, recursive=True)

        # Process each file
        for file in matching_files:
            if file in found_files:
                logging.debug(f"duplicate file: {file}")
                continue

            logging.debug(f"Adding file: {file}")

            found_files.add(file)

            # Get absolute path
            file_path = os.path.abspath(file)

            driver_prop = ZephyrDriver(filepath=file_path)

            # Skip if we only want files with compat values
            if compat_only and not driver_prop.dt_compats:
                continue

            try:
                if zephyr_def:
                    driver_prop.load_zephyr_def(zephyr_def.get(driver_prop.api_struct, []))
            except Exception as e:
                logging.debug(f"\nError driver_prop {driver_prop}: {e}")

            # Add to our entries list
            entries.append(driver_prop)

    return entries


def export_to_xlsx(output_xlsx, driver_list):
    #
    # create a spreadsheet where each tab is a unique driver class output:
    #    driver_class, filename, dt_compat, PM, API structure, api1, api2, api3...
    # where:
    #    each API is either true/false. Later, someone can edit the spreadsheet and
    #    use N/A if the API function doesn't make sense for the IP
    #

    # Get unique list of classes found in driver_list
    class_groups = defaultdict(list)
    for driver_item in driver_list:
        key = driver_item.driver_class
        class_groups[key].append(driver_item)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    # yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    with pd.ExcelWriter(output_xlsx, mode="w", engine="openpyxl") as writer:
        for class_key, objs in sorted(
            class_groups.items(),
            key=lambda kv: (kv[0] is not None, "" if kv[0] is None else kv[0].casefold()),
        ):
            df = pd.DataFrame([o.api_list() for o in objs])
            sheet = "None" if class_key is None else str(class_key)[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.sheets[sheet]
            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        value = cell.value
                        if isinstance(value, bool):
                            value = "TRUE" if value else "FALSE"
                        if value:
                            max_length = max(max_length, len(str(value)))
                    except (AttributeError, TypeError):
                        pass
                # add 4 so text isn't right against the boarder
                ws.column_dimensions[column].width = max_length + 6

            # Freeze first row and turn on filters for all columns
            ws.freeze_panes = "B2"
            ws.auto_filter.ref = ws.dimensions

            data_range = f"A2:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
            ws.conditional_formatting.add(
                data_range, FormulaRule(formula=['ISNUMBER(SEARCH("TRUE",A2))'], fill=green)
            )
            ws.conditional_formatting.add(
                data_range, FormulaRule(formula=['ISNUMBER(SEARCH("FALSE",A2))'], fill=red)
            )


def main():
    parser = argparse.ArgumentParser(
        description="Scrape driver properties",
        allow_abbrev=False,  # <-- This disables abbreviations
    )

    parser.add_argument(
        'patterns', nargs='+', help='One or more wildcard patterns (e.g., "drivers/*/*mcux*.c")'
    )

    parser.add_argument(
        '-i',
        '--include',
        type=Path,
        help='Scan for Zephyr __subsystem API structs (zephyr/include)',
    )

    parser.add_argument(
        '-r', '--recursive', action='store_true', help='Force recursive search for patterns'
    )

    parser.add_argument(
        '-c',
        '--compat-only',
        action='store_true',
        help='Include only files with DT_DRV_COMPAT values',
    )

    parser.add_argument('-x', '--xlsx', help='Output data to xlsx file')

    parser.add_argument(
        '-v',
        '--verbose',
        action="count",
        default=0,
        help="Increase log verbosity (-v for INFO, -vv for DEBUG).",
    )

    args = parser.parse_args()

    configure_logging(args.verbose)
    logging.info("Starting driver processing...")
    logging.debug("Arguments: %s", args)

    zephyr_def = None
    if args.include:
        zephyr_def = scan_api_structs(args.include)
        logging.info(f"Scanning API information from {args.include}")

    driver_list = find_driver_files(
        args.patterns, compat_only=args.compat_only, recursive=args.recursive, zephyr_def=zephyr_def
    )

    if args.xlsx:
        export_to_xlsx(args.xlsx, driver_list)


if __name__ == "__main__":
    main()
